from pathlib import Path
import re
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "data" / "bible_key_verses.xlsx"
CANONICAL_COUNTS = {'Genesis': 50, 'Exodus': 40, 'Leviticus': 27, 'Numbers': 36, 'Deuteronomy': 34, 'Joshua': 24, 'Judges': 21, 'Ruth': 4, '1 Samuel': 31, '2 Samuel': 24, '1 Kings': 22, '2 Kings': 25, '1 Chronicles': 29, '2 Chronicles': 36, 'Ezra': 10, 'Nehemiah': 13, 'Esther': 10, 'Job': 42, 'Psalms': 150, 'Proverbs': 31, 'Ecclesiastes': 12, 'Song of Solomon': 8, 'Isaiah': 66, 'Jeremiah': 52, 'Lamentations': 5, 'Ezekiel': 48, 'Daniel': 12, 'Hosea': 14, 'Joel': 3, 'Amos': 9, 'Obadiah': 1, 'Jonah': 4, 'Micah': 7, 'Nahum': 3, 'Habakkuk': 3, 'Zephaniah': 3, 'Haggai': 2, 'Zechariah': 14, 'Malachi': 4, 'Matthew': 28, 'Mark': 16, 'Luke': 24, 'John': 21, 'Acts': 28, 'Romans': 16, '1 Corinthians': 16, '2 Corinthians': 13, 'Galatians': 6, 'Ephesians': 6, 'Philippians': 4, 'Colossians': 4, '1 Thessalonians': 5, '2 Thessalonians': 3, '1 Timothy': 6, '2 Timothy': 4, 'Titus': 3, 'Philemon': 1, 'Hebrews': 13, 'James': 5, '1 Peter': 5, '2 Peter': 3, '1 John': 5, '2 John': 1, '3 John': 1, 'Jude': 1, 'Revelation': 22}
CANONICAL_ORDER = list(CANONICAL_COUNTS)

def slug(name):
    return re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")

def load_data():
    wb = load_workbook(WORKBOOK, data_only=True)
    bws = wb["Books"]
    cws = wb["Chapters_Data"]
    book_rows = {}
    for row in bws.iter_rows(min_row=3, values_only=True):
        if not row[0]: continue
        book_rows[str(row[0]).strip()] = {
            "book": str(row[0]).strip(), "testament": row[1], "genre": row[2],
            "chapters": int(row[3]), "summary": str(row[4] or "").strip()
        }
    chapter_rows=[]
    for row in cws.iter_rows(min_row=2, max_col=9, values_only=True):
        if not row[0]: continue
        chapter_rows.append({
            "book": str(row[0]).strip(), "chapter": int(row[1]),
            "verse": str(row[2] or "").strip(), "identifier": str(row[3] or "").strip(),
            "why": str(row[4] or "").strip(), "insight": str(row[5] or "").strip(),
            "major_characters": str(row[6] or "").strip(), "keywords": str(row[7] or "").strip(),
            "canonical_section": str(row[8] or "").strip()
        })
    return book_rows, chapter_rows
